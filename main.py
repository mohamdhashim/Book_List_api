from flask import Flask
from flask_restful import reqparse, abort, Api, Resource
from openpyxl import load_workbook  # to manipulating our excel sheets


app = Flask(__name__)
api = Api(app)

# First json Design
# books = [
#     {
#         'order': 1,
#         'book': {'name': 'Cairo trilogy',
#                  'url': 'www.wiki/trilogy.com'},

#         'author': {name': 'nagib mahfouz',
#                    'url': 'www.wiki/nagib_mahfouz.com'},

#         'country': {'name': 'egypt',
#                     'url': 'www.wiki/egypt.com'}
#     }
# ]


wb = load_workbook('booklist.xlsx')
ws = wb['Sheet']


def abort_if_novel_doesnt_exist(order):
    abort(404, message="Novel {} doesn't exist".format(order))


parser = reqparse.RequestParser()

parser.add_argument('order', help="Order cannot be blank!")

parser.add_argument('novel_name', required=True,
                    help="Novel name cannot be blank!")
parser.add_argument('novel_url')

parser.add_argument('author_name', required=True,
                    help="Author name cannot be blank!")
parser.add_argument('author_url')

parser.add_argument('country_name', required=True,
                    help="Country name cannot be blank!")
parser.add_argument('country_url')


def row_to_json(row):
    # convert excel row to json format
    json = {
        'order': row[0].value,
        'novel_name': row[1].value,
        'novel_url': row[1].hyperlink.target if row[1].hyperlink else None,

        'author_name': row[2].value,
        'author_url': row[2].hyperlink.target if row[1].hyperlink else None,

        'country_name':  row[3].value,
        'country_url': row[3].hyperlink.target if row[1].hyperlink else None,
    }
    return json


def reorder_sheet(order, up=True):
    # reorder the sheet after any changes
    # up = True ==> to shuffle up rows ex rows[25:n] ==> rows[24:n-1] (ex : delete row)
    # up = Flase ==> to shuffle down rows ex rows[25:n] ==> rows[26:n+1] (ex : add row in the middle)

    if up:
        for i in range(2, ws.max_row+1):
            if ws.cell(i, 1).value and ws.cell(i, 1).value >= int(order):
                ws.cell(i, 1).value = ws.cell(i, 1).value-1
    else:
        for i in range(2, ws.max_row+1):
            if ws.cell(i, 1).value and ws.cell(i, 1).value >= int(order):
                ws.cell(i, 1).value = ws.cell(i, 1).value+1


# Novel
# shows a single Novel and lets you delete a Novel or Edit it
class Novel(Resource):
    
    def get(self, order):
        # Get by Novel Order
        for row in ws.iter_rows(): 
            if str(row[0].value) == order:
                return row_to_json(row)

        abort_if_novel_doesnt_exist(order)


    def delete(self, order):
        '''
            deleting row/novel by it's order and reorgranize Sheet
            to fill the deleted row empty place by shiffting the following rows ==> row-1
        '''
        
        deleted = False
        for i in range(1, ws.max_row+1):
            if ws._get_cell(i, 1).value == int(order):
                deleted = True

            #shifting all bellow range of rows to start from the empty row of the deleted
            if deleted: 
                for j in range(1, ws.max_column+1):
                    ws.cell(i, j).value = ws.cell(i+1, j).value
                    ws.cell(i, j).hyperlink = ws.cell(i+1, j).hyperlink

        if not deleted:
            abort_if_novel_doesnt_exist(order)

        reorder_sheet(order)
        wb.save('booklist.xlsx')
        return 'deleted successfully', 204


    def put(self, order):
        args = parser.parse_args()
        for i in range(2, ws.max_row+1):
            if ws.cell(i, 1).value == int(order):

                ws.cell(i, 2).value = args['novel_name']
                ws.cell(i, 2).hyperlink = args['novel_url']

                ws.cell(i, 3).value = args['author_name']
                ws.cell(i, 3).hyperlink = args['author_url']

                ws.cell(i, 4).value = args['country_name']
                ws.cell(i, 4).hyperlink = args['country_url']
                wb.save('booklist.xlsx')
                return row_to_json(ws[i])

        abort_if_novel_doesnt_exist(order)


# BookList
# shows a list of all books, and lets you POST to add new books
class NovelList(Resource):

    def get(self):

        books = []  # json (list of dictionaries)
        for i, book in enumerate(ws.iter_rows()):
            if i == 0:
                continue  # the row of the head
            books.append(row_to_json(book))

        return books

    def post(self):
        ''' 
            insert new novel with new order
            all the rows from the same position will be move to next row
            ex: insert['order=25'] ==> all the rows order from rows[25:row_max] will 
            increase +1 ==> rows[26: row_max+1] 
        '''
        args = parser.parse_args()
        order = int(args['order'])

        # insert(order = None) will append our novels table
        if(not order or order > ws.max_row):
            index = ws.max_row+1
            order = ws.max_row
        else:
            index = order+1 # body + head
            
        reorder_sheet(order, up=False)

        ws.insert_rows(index)
        
        if(not order or order > ws.max_row): 
            order = order - 1
            
        ws.cell(index, 1).value = order 
        ws.cell(index, 2).value = args['novel_name']
        ws.cell(index, 2).hyperlink = args['novel_url']

        ws.cell(index, 3).value = args['author_name']
        ws.cell(index, 3).hyperlink = args['author_url']

        ws.cell(index, 4).value = args['country_name']
        ws.cell(index, 4).hyperlink = args['country_url']

        wb.save('booklist.xlsx')
        return row_to_json(ws[index]), 201


##
# Actually setup the Api resource routing here
##
api.add_resource(NovelList, '/novels')
api.add_resource(Novel, '/novels/<order>')

if __name__ == '__main__':
    app.run()
